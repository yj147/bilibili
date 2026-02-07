#!/usr/bin/env python3
"""
Unified Scraper Template (v3.0) - Production Grade

A production-ready web scraping template incorporating battle-tested patterns:
- Semaphore-based concurrency control
- Browser context reuse with stealth injection
- Proxy support (HTTP/HTTPS with auth)
- Cache-first deduplication
- Cookie injection for authenticated sessions
- Pagination support with depth limits
- Multiple storage backends (JSONL, JSON, CSV)
- Rate limiting with random delays
- Graceful shutdown with partial result preservation

This is a TEMPLATE - override the `parse()` method with your extraction logic.
"""

import asyncio
import yaml
import logging
import random
import json
import csv
import signal
from typing import Dict, Any, List, Optional, Set
from pathlib import Path
from dataclasses import dataclass
from abc import ABC, abstractmethod

# --- Dependency Checks ---
try:
    from curl_cffi.requests import AsyncSession as CurlAsyncSession
except ImportError:
    print("ERROR: pip install 'curl_cffi>=0.6.0'")
    exit(1)

try:
    from playwright.async_api import async_playwright, Browser, BrowserContext
except ImportError:
    print("ERROR: pip install playwright && playwright install chromium")
    exit(1)

try:
    from bs4 import BeautifulSoup
except ImportError:
    print("ERROR: pip install beautifulsoup4 lxml")
    exit(1)

# --- Stealth JS (fallback when playwright-stealth not installed) ---
STEALTH_JS = """
Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
delete navigator.__proto__.webdriver;
window.chrome = {
    runtime: { id: 'mocked' },
    loadTimes: function() { return {}; },
    csi: function() { return {}; },
    app: { isInstalled: false },
};
Object.defineProperty(navigator, 'plugins', {
    get: () => [
        { name: 'Chrome PDF Plugin', filename: 'internal-pdf-viewer' },
        { name: 'Chrome PDF Viewer', filename: 'mhjfbmdgcfjbbpaeojofohoefgiehjai' },
        { name: 'Native Client', filename: 'internal-nacl-plugin' },
    ],
});
Object.defineProperty(navigator, 'languages', {
    get: () => ['zh-CN', 'zh', 'en-US', 'en'],
});
"""

# --- Logging Setup ---
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger("UnifiedScraper")


# --- Custom Exceptions ---
class ScraperError(Exception):
    """Base exception for scraper errors."""

class IPBlockError(ScraperError):
    """Server blocked our IP (e.g., 403, 429, or platform-specific codes)."""

class DataFetchError(ScraperError):
    """Failed to fetch or parse data from API/page."""

class ContentNotFoundError(ScraperError):
    """Requested content does not exist (no retry needed)."""


# --- Configuration Loader ---
def load_config(path: str = "config.yaml") -> Dict[str, Any]:
    """Load and validate YAML configuration."""
    config_path = Path(path)
    if not config_path.exists():
        raise FileNotFoundError(f"Config not found: {config_path}")
    with open(config_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


# --- Storage Backend Abstraction ---
class AbstractStorage(ABC):
    """Abstract base class for storage backends."""

    @abstractmethod
    def save(self, results: List[Dict[str, Any]], filepath: Path) -> None:
        """Save results to storage."""
        pass


class JSONLStorage(AbstractStorage):
    """JSONL storage backend (one JSON object per line)."""

    def save(self, results: List[Dict[str, Any]], filepath: Path) -> None:
        filepath.parent.mkdir(parents=True, exist_ok=True)
        with open(filepath, "w", encoding="utf-8") as f:
            for item in results:
                f.write(json.dumps(item, ensure_ascii=False) + "\n")


class JSONStorage(AbstractStorage):
    """JSON storage backend (single JSON array)."""

    def save(self, results: List[Dict[str, Any]], filepath: Path) -> None:
        filepath.parent.mkdir(parents=True, exist_ok=True)
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(results, f, ensure_ascii=False, indent=2)


class CSVStorage(AbstractStorage):
    """CSV storage backend."""

    def save(self, results: List[Dict[str, Any]], filepath: Path) -> None:
        if not results:
            return
        filepath.parent.mkdir(parents=True, exist_ok=True)
        keys = results[0].keys()
        with open(filepath, "w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=keys)
            writer.writeheader()
            writer.writerows(results)


class SQLiteStorage(AbstractStorage):
    """SQLite storage backend for large datasets."""

    def save(self, results: List[Dict[str, Any]], filepath: Path) -> None:
        if not results:
            return
        import sqlite3
        filepath.parent.mkdir(parents=True, exist_ok=True)
        db_path = filepath.with_suffix(".db")
        conn = sqlite3.connect(str(db_path))
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=NORMAL")

        keys = list(results[0].keys())
        col_defs = ", ".join(f'"{k}" TEXT' for k in keys)
        conn.execute(f'CREATE TABLE IF NOT EXISTS scraped_data (id INTEGER PRIMARY KEY, {col_defs})')

        placeholders = ", ".join(["?"] * len(keys))
        col_names = ", ".join(f'"{k}"' for k in keys)
        batch_size = 5000
        for i in range(0, len(results), batch_size):
            batch = results[i:i + batch_size]
            rows = [tuple(str(item.get(k, "")) for k in keys) for item in batch]
            conn.executemany(f'INSERT INTO scraped_data ({col_names}) VALUES ({placeholders})', rows)
            conn.commit()

        conn.close()
        logger.info(f"SQLite: {len(results)} rows → {db_path}")


class ExcelStorage(AbstractStorage):
    """Excel storage backend (requires openpyxl)."""

    def save(self, results: List[Dict[str, Any]], filepath: Path) -> None:
        if not results:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError:
            logger.error("Excel export requires: pip install openpyxl")
            return

        filepath = filepath.with_suffix(".xlsx")
        filepath.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "Scraped Data"

        headers = list(results[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        for row_idx, item in enumerate(results, 2):
            for col_idx, key in enumerate(headers, 1):
                value = item.get(key, "")
                if isinstance(value, list):
                    value = ", ".join(str(v) for v in value)
                ws.cell(row=row_idx, column=col_idx, value=str(value))

        wb.save(str(filepath))
        logger.info(f"Excel: {len(results)} rows → {filepath}")


class StorageFactory:
    """Factory for creating storage backends."""

    BACKENDS = {
        "jsonl": JSONLStorage,
        "json": JSONStorage,
        "csv": CSVStorage,
        "sqlite": SQLiteStorage,
        "excel": ExcelStorage,
    }

    @staticmethod
    def create(storage_type: str) -> AbstractStorage:
        backend_class = StorageFactory.BACKENDS.get(storage_type)
        if not backend_class:
            raise ValueError(f"Unknown storage type: {storage_type}")
        return backend_class()


# --- Cache Manager ---
class CacheManager:
    """Simple JSON file-based cache for URL deduplication."""

    def __init__(self, filepath: Optional[Path] = None, enabled: bool = True):
        self.enabled = enabled
        self.filepath = filepath
        self.cache: Set[str] = set()
        if self.enabled and self.filepath:
            self._load()

    def _load(self) -> None:
        """Load cache from disk."""
        if self.filepath and self.filepath.exists():
            try:
                with open(self.filepath, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    self.cache = set(data.get("urls", []))
                logger.info(f"Loaded {len(self.cache)} URLs from cache")
            except Exception as e:
                logger.warning(f"Failed to load cache: {e}")

    def _save(self) -> None:
        """Save cache to disk."""
        if not self.enabled or not self.filepath:
            return
        try:
            self.filepath.parent.mkdir(parents=True, exist_ok=True)
            with open(self.filepath, "w", encoding="utf-8") as f:
                json.dump({"urls": list(self.cache)}, f, indent=2)
        except Exception as e:
            logger.error(f"Failed to save cache: {e}")

    def has(self, url: str) -> bool:
        """Check if URL is in cache."""
        return self.enabled and url in self.cache

    def add(self, url: str) -> None:
        """Add URL to cache."""
        if self.enabled:
            self.cache.add(url)

    def save(self) -> None:
        """Public method to save cache."""
        self._save()


# --- Main Scraper Class ---
class UnifiedScraper:
    """Production-grade configuration-driven web scraper."""

    def __init__(self, config: Dict[str, Any]):
        self.config = config
        self.strategy = config.get("crawl_strategy", {})
        self.anti_detection = config.get("anti_detection", {})
        self.auth_config = config.get("auth", {})
        self.cache_config = config.get("cache", {})
        self.rules = config.get("extraction_rules", {})
        self.storage_config = config.get("storage", {})

        # State
        self.results: List[Dict[str, Any]] = []
        self.queue: asyncio.Queue = asyncio.Queue()
        self.visited: Set[str] = set()
        self.browser: Optional[Browser] = None
        self.browser_context: Optional[BrowserContext] = None
        self._playwright = None
        self.shutdown_requested = False
        self._save_interval = self.storage_config.get("save_interval", 50)
        self._unsaved_count = 0

        # Cache manager
        cache_enabled = self.cache_config.get("enabled", False)
        cache_path = self.cache_config.get("filepath")
        self.cache = CacheManager(
            filepath=Path(cache_path) if cache_path else None,
            enabled=cache_enabled
        )

        # Storage backend
        storage_type = self.storage_config.get("type", "jsonl")
        self.storage = StorageFactory.create(storage_type)

        # Setup signal handlers for graceful shutdown
        signal.signal(signal.SIGINT, self._signal_handler)
        signal.signal(signal.SIGTERM, self._signal_handler)

    def _signal_handler(self, signum, frame):
        """Handle shutdown signals gracefully."""
        logger.warning("Shutdown signal received, finishing current tasks...")
        self.shutdown_requested = True

    async def run(self):
        """Main entry point to start the scraping process."""
        try:
            targets = self.config.get("targets", [])
            if not targets:
                logger.warning("No targets defined in configuration")
                return

            # Initialize browser if needed
            mode = self.strategy.get("mode", "http")
            if mode == "browser":
                await self._init_browser()

            # Process all targets
            for target in targets:
                if self.shutdown_requested:
                    break
                await self._crawl_target(target)

        except Exception as e:
            logger.error(f"Critical error: {e}", exc_info=True)
        finally:
            await self._cleanup()

    async def _init_browser(self):
        """Initialize browser context once and reuse."""
        self._playwright = await async_playwright().start()

        # Proxy configuration
        proxy_config = None
        proxy_settings = self.anti_detection.get("proxy", {})
        if proxy_settings.get("enabled") and proxy_settings.get("url"):
            proxy_config = {"server": proxy_settings["url"]}

        headless = self.strategy.get("headless", True)
        user_data_dir = self.strategy.get("user_data_dir", "")

        context_options = {
            "viewport": {"width": 1920, "height": 1080},
            "user_agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "locale": "zh-CN",
            "args": [
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox",
                "--disable-dev-shm-usage",
            ],
        }
        if proxy_config:
            context_options["proxy"] = proxy_config

        if user_data_dir:
            # Persistent context: keeps login state, cookies across runs
            self.browser_context = await self._playwright.chromium.launch_persistent_context(
                user_data_dir=user_data_dir,
                headless=headless,
                **context_options,
            )
        else:
            self.browser = await self._playwright.chromium.launch(headless=headless)
            self.browser_context = await self.browser.new_context(**context_options)

        # Inject cookies if configured
        cookies_str = self.auth_config.get("cookies", "")
        if cookies_str:
            await self._inject_cookies(cookies_str)

        # Inject stealth scripts
        try:
            from playwright_stealth import stealth_async
            page = await self.browser_context.new_page()
            await stealth_async(page)
            await page.close()
        except ImportError:
            # Fallback: inject inline stealth JS via init_script
            stealth_path = self.anti_detection.get("stealth_js_path")
            if stealth_path and Path(stealth_path).exists():
                await self.browser_context.add_init_script(path=stealth_path)
            else:
                await self.browser_context.add_init_script(STEALTH_JS)
            logger.info("Using built-in stealth JS (playwright-stealth not installed)")

        logger.info("Browser context initialized and ready")

    async def _inject_cookies(self, cookies_str: str):
        """Parse and inject cookies into browser context."""
        if not self.browser_context:
            return

        # Derive domain from first target's allowed_domains, or fall back to auth config
        targets = self.config.get("targets", [])
        domain = self.auth_config.get("cookie_domain", "")
        if not domain and targets:
            allowed = targets[0].get("allowed_domains", [])
            if allowed:
                domain = f".{allowed[0]}" if not allowed[0].startswith(".") else allowed[0]

        if not domain:
            logger.warning("No cookie domain configured, cookies may not be sent correctly")
            domain = "localhost"

        cookies = []
        for pair in cookies_str.split(";"):
            pair = pair.strip()
            if "=" in pair:
                key, value = pair.split("=", 1)
                cookies.append({
                    "name": key.strip(),
                    "value": value.strip(),
                    "domain": domain,
                    "path": "/"
                })

        if cookies:
            await self.browser_context.add_cookies(cookies)
            logger.info(f"Injected {len(cookies)} cookies for domain {domain}")

    async def _crawl_target(self, target: Dict[str, Any]):
        """Crawl a single target with all its start URLs."""
        start_urls = target.get("start_urls", [])
        depth_limit = target.get("depth_limit", 0)

        logger.info(f"Starting crawl: '{target.get('name')}' ({len(start_urls)} URLs)")

        # Enqueue start URLs
        for url in start_urls:
            await self.queue.put((url, 0))  # (url, depth)

        # Semaphore for concurrency control
        concurrency = self.strategy.get("concurrency", 5)
        semaphore = asyncio.Semaphore(concurrency)

        # Worker tasks
        workers = [
            asyncio.create_task(self._worker(semaphore, depth_limit))
            for _ in range(concurrency)
        ]

        # Wait for queue to be empty
        await self.queue.join()

        # Cancel workers
        for worker in workers:
            worker.cancel()

        await asyncio.gather(*workers, return_exceptions=True)

    async def _worker(self, semaphore: asyncio.Semaphore, depth_limit: int):
        """Worker coroutine that processes URLs from the queue."""
        while not self.shutdown_requested:
            try:
                url, depth = await asyncio.wait_for(self.queue.get(), timeout=1.0)
            except asyncio.TimeoutError:
                continue

            try:
                # Skip if already visited or cached
                if url in self.visited or self.cache.has(url):
                    logger.debug(f"Skipping (cached/visited): {url}")
                    self.queue.task_done()
                    continue

                self.visited.add(url)

                # Process URL with semaphore
                async with semaphore:
                    await self._process_url(url, depth, depth_limit)

                    # Rate limiting
                    delay_min = self.strategy.get("delay", {}).get("min", 1.0)
                    delay_max = self.strategy.get("delay", {}).get("max", 2.5)
                    await asyncio.sleep(random.uniform(delay_min, delay_max))

            except Exception as e:
                logger.error(f"Worker error processing {url}: {e}")
            finally:
                self.queue.task_done()

    async def _process_url(self, url: str, depth: int, depth_limit: int):
        """Fetch, parse, and handle a single URL."""
        html = await self._fetch(url)
        if not html:
            return

        # Parse content
        data = self.parse(html, url)
        if data:
            self.results.append(data)
            self.cache.add(url)
            self._unsaved_count += 1
            logger.info(f"✓ Parsed: {url}")

            # Incremental save to prevent data loss on crash
            if self._save_interval > 0 and self._unsaved_count >= self._save_interval:
                filepath = Path(self.storage_config.get("filepath", "output/data.jsonl"))
                self.storage.save(self.results, filepath)
                self.cache.save()
                self._unsaved_count = 0
                logger.info(f"Checkpoint: saved {len(self.results)} items")

        # Handle pagination
        if depth_limit == 0 or depth < depth_limit:
            next_urls = self._extract_next_pages(html)
            for next_url in next_urls:
                if next_url not in self.visited:
                    await self.queue.put((next_url, depth + 1))

    def _extract_next_pages(self, html: str) -> List[str]:
        """Extract pagination URLs from HTML."""
        next_page_selector = self.rules.get("listing_page", {}).get("next_page_selector")
        if not next_page_selector:
            return []

        soup = BeautifulSoup(html, "lxml")
        # Simple CSS selector parsing (user should customize)
        next_links = soup.select(next_page_selector)
        return [link.get("href") for link in next_links if link.get("href")]

    async def _fetch(self, url: str) -> Optional[str]:
        """Fetch URL using configured strategy with retries and error classification."""
        mode = self.strategy.get("mode", "http")
        max_retries = self.strategy.get("max_retries", 3)

        for attempt in range(max_retries):
            try:
                if mode == "browser":
                    return await self._fetch_with_browser(url)
                else:
                    return await self._fetch_with_http(url)
            except ContentNotFoundError:
                logger.info(f"Content not found (no retry): {url}")
                return None
            except IPBlockError as e:
                logger.error(f"IP blocked on {url}: {e}")
                # Longer backoff for IP blocks
                if attempt < max_retries - 1:
                    await asyncio.sleep(5 * (2 ** attempt))
            except Exception as e:
                logger.warning(f"Attempt {attempt + 1}/{max_retries} failed for {url}: {e}")
                if attempt < max_retries - 1:
                    await asyncio.sleep(2 ** attempt)

        return None

    async def _fetch_with_http(self, url: str) -> str:
        """Fast HTTP fetching with curl-cffi."""
        impersonate = self.anti_detection.get("impersonate", "chrome110")
        timeout = self.strategy.get("timeout", 30)

        # Proxy configuration
        proxies = None
        proxy_settings = self.anti_detection.get("proxy", {})
        if proxy_settings.get("enabled") and proxy_settings.get("url"):
            proxies = {"http": proxy_settings["url"], "https": proxy_settings["url"]}

        # Cookie headers
        headers = {}
        cookies_str = self.auth_config.get("cookies", "")
        if cookies_str:
            headers["Cookie"] = cookies_str

        async with CurlAsyncSession() as session:
            resp = await session.get(
                url,
                impersonate=impersonate,
                timeout=timeout,
                proxies=proxies,
                headers=headers if headers else None,
            )
            if resp.status_code in (403, 429):
                raise IPBlockError(f"HTTP {resp.status_code} — likely IP blocked or rate limited")
            if resp.status_code == 404:
                raise ContentNotFoundError(f"HTTP 404 for {url}")
            resp.raise_for_status()
            return resp.text

    async def _fetch_with_browser(self, url: str) -> str:
        """Browser-based fetching with reused context."""
        if not self.browser_context:
            raise RuntimeError("Browser context not initialized")

        page = await self.browser_context.new_page()
        try:
            timeout = self.strategy.get("timeout", 30) * 1000
            await page.goto(url, wait_until="networkidle", timeout=timeout)
            content = await page.content()
            return content
        finally:
            await page.close()

    def parse(self, html: str, url: str) -> Optional[Dict[str, Any]]:
        """
        Parse HTML to extract structured data.

        *** OVERRIDE THIS METHOD WITH YOUR EXTRACTION LOGIC ***

        This is a placeholder. Use BeautifulSoup, self.rules, and your
        domain knowledge to extract the data you need.
        """
        soup = BeautifulSoup(html, "lxml")
        title = soup.title.string.strip() if soup.title else "No Title"

        return {
            "url": url,
            "title": title,
            "content_length": len(html),
        }

    async def _cleanup(self):
        """Cleanup resources and save results."""
        logger.info("Cleaning up...")

        # Save results
        if self.results:
            filepath = Path(self.storage_config.get("filepath", "output/data.jsonl"))
            self.storage.save(self.results, filepath)
            logger.info(f"Saved {len(self.results)} items to {filepath}")

        # Save cache
        self.cache.save()

        # Close browser
        if self.browser_context:
            await self.browser_context.close()
        if self.browser:
            await self.browser.close()

        logger.info("Cleanup complete")


# --- Main Execution ---
async def main():
    """Main async function to run the scraper."""
    try:
        config = load_config("config.yaml")
        scraper = UnifiedScraper(config)
        await scraper.run()
    except FileNotFoundError as e:
        logger.error(f"Configuration error: {e}")
    except KeyboardInterrupt:
        logger.warning("Interrupted by user")
    except Exception as e:
        logger.critical(f"Fatal error: {e}", exc_info=True)


if __name__ == "__main__":
    asyncio.run(main())
